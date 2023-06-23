import os
import sys
import pandas as pd
import numpy as np
import json
import cv2
import hashlib
from werkzeug.utils import secure_filename
from bs4 import BeautifulSoup
import openpyxl

__dir__ = os.path.dirname(os.path.abspath(__file__))
sys.path.append(__dir__)
sys.path.insert(0, os.path.abspath(os.path.join(__dir__, "..")))
sys.path.insert(0, os.path.abspath(os.path.join(__dir__, "..\PaddleOCR")))
sys.path.insert(0, os.path.abspath(os.path.join(__dir__, "..\PaddleOCR\ppstructure")))
sys.path.insert(
    0, os.path.abspath(os.path.join(__dir__, "..\PaddleOCR\ppstructure\table"))
)

from config import ALLOWED_EXTENSIONS, EXTRACT_RESULT_PATH, JSON_DIR, UPLOAD_IMAGE_PATH
from PaddleOCR import PPStructure, save_structure_res
from repository import tableRepo, documentTypeRepo


table_repo = tableRepo.TableRepo()
doc_type_repo = documentTypeRepo.DocumentTypeRepo()


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def get_hash_code(img_path):
    # Read the file in binary mode and calculate the MD5 hash
    with open(img_path, "rb") as file:
        content = file.read()
        md5_hash = hashlib.md5(content).hexdigest()
    return md5_hash


def is_unique(hash_code):
    result = table_repo.get_table_extraction_by_hash_code(hash_code)
    return result is None


# return : excel file path
def extract_table(img_path):
    table_engine = PPStructure(layout=False, show_log=True)
    save_folder = EXTRACT_RESULT_PATH
    img = cv2.imread(img_path)
    result = table_engine(img)
    img_filename = os.path.basename(img_path).split(".")[0]
    excel_path = save_structure_res(result, save_folder, img_filename)

    return excel_path


def extract_html(img_path):
    table_engine = PPStructure(layout=False, show_log=True)
    img = cv2.imread(img_path)
    result = table_engine(img)
    html_path = os.path.join(EXTRACT_RESULT_PATH, "show.html")
    f_html = open(html_path, mode="w", encoding="utf-8")
    pred_html = result[0]["res"]["html"]
    f_html.write(
        '<table  border="1">'
        + pred_html.replace("<html><body><table>", "").replace(
            "</table></body></html>", ""
        )
    )
    f_html.close()

    return html_path


#####


def header_handler(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    sheet_names = wb.sheetnames
    sheet = wb[sheet_names[0]]
    merged_cells = sheet.merged_cells
    if merged_cells:
        last_row_number = get_header_last_row_number(merged_cells)
    else:
        last_row_number = 1
    header_list = get_header_list(sheet)
    json_data = {}
    json_data["header_list"] = header_list
    json_data["skiprows"] = last_row_number
    return json_data


def get_header_last_row_number(merged_cells):
    merged_list = []

    for merged_cell in merged_cells:
        merged_list.append(str(merged_cell))
    print(merged_list)
    rowspan_list = []

    for merged in merged_list:
        start_cell, end_cell = merged.split(":")
        if start_cell[0] == end_cell[0]:
            rowspan_list.append(merged.split(":"))
    print(rowspan_list)
    merged_row_numbers = []
    for rowspan in rowspan_list:
        if rowspan[0][1] == "1":
            merged_row_numbers.append(int(rowspan[1][1]))

    merged_row_numbers.sort(reverse=True)
    return merged_row_numbers[0]


def get_header_list(sheet):
    print("get header list")
    merged_cells_ranges = sheet.merged_cells.ranges
    print(merged_cells_ranges)
    sorted_ranges = sorted(merged_cells_ranges, key=lambda x: x.min_col)
    print(sorted_ranges)
    header_list = []
    if sorted_ranges:
        for merged_range in sorted_ranges:
            if merged_range.min_col == merged_range.max_col:
                header_list.append(merged_range.start_cell.value)
                print(merged_range.start_cell.value)
            else:
                for i in range(
                    merged_range.min_col, merged_range.max_col + 1
                ):  # 3 kali
                    val = str(merged_range.start_cell.value)
                    val = (
                        val
                        + " - "
                        + str(sheet.cell(int(merged_range.min_row) + 1, i).value)
                    )
                    print(val)
                    header_list.append(val)
    else:
        for cell in sheet.iter_cols(min_row=1, max_row=1, values_only=True):
            header_list.append(cell[0])
    return header_list


#####


def merge_table_header(html_table):
    # Membaca tabel HTML menggunakan BeautifulSoup
    soup = BeautifulSoup(html_table, "html.parser")

    # Temukan semua elemen tr dalam tabel
    rows = soup.find_all("tr")

    # Mencari elemen header dan data
    row_header = []
    row_data = []
    for row in rows:
        if row.get("class") and "header" in row.get("class"):
            row_header.append(row)
        else:
            row_data.append(str(row))

    # Jika hanya ada satu baris header, kembalikan tabel HTML yang ada
    if len(row_header) == 1:
        return str(soup)

    # Menggabungkan header jika ada dua baris header
    if len(row_header) == 2:
        new_header = []

        # Membaca header baris pertama
        header_row_1 = row_header[0].find_all("td")
        for cell in header_row_1:
            rowspan = cell.get("rowspan", 1)
            colspan = cell.get("colspan", 1)

            if int(rowspan) > 1:
                new_th = soup.new_tag("th")
                new_th.string = cell.text.strip()
                new_header.append(str(new_th))

            if int(colspan) > 1:
                # Membaca header pada baris berikutnya
                header_row_2 = row_header[1].find_all("td")

                for i in range(int(colspan)):
                    new_th = soup.new_tag("th")
                    new_th.string = (
                        f"{cell.text.strip()} - {header_row_2[i].text.strip()}"
                    )
                    new_header.append(str(new_th))

    return (new_header, row_data)


def generate_new_table(list_header, list_row):
    # Membuat BeautifulSoup objek dari string HTML kosong
    soup = BeautifulSoup("", "html.parser")

    # Membuat elemen <table> baru
    table = soup.new_tag("table")
    print(table)

    # Menambahkan elemen header ke dalam tabel
    header_row = soup.new_tag("tr")
    for header_cell in list_header:
        print(header_cell)
        header_row.append(BeautifulSoup(header_cell, "html.parser"))
        print(header_row)
    table.append(header_row)
    print("debug 2")
    # Menambahkan elemen data ke dalam tabel
    for data_row in list_row:
        table.append(BeautifulSoup(data_row, "html.parser"))

    # Menggabungkan semua elemen menjadi satu
    soup.append(table)
    print("debug 3")

    # Menghasilkan tabel HTML baru
    new_html_table = str(soup)

    return new_html_table


def html_to_json(html):
    soup = BeautifulSoup(html, "html.parser")

    table_data = []
    table = soup.find("table")
    table_headings = [th.text for th in table.find_all("th")]

    for row in table.find_all("tr")[1:]:
        row_data = [td.text for td in row.find_all("td")]
        table_data.append(dict(zip(table_headings, row_data)))

    json_data = json.dumps(table_data, indent=4)
    return json_data


def html_transform(html_str):
    headers, row_data = merge_table_header(html_str)
    new_html = generate_new_table(headers, row_data)
    json = html_to_json(new_html)
    return json


def data_transform(document_type, excel_path, hash_code):
    document_type = document_type.replace("_", " ").title()
    dokumen = doc_type_repo.get_doc_type_by_name(document_type)
    if dokumen["jenis_tabel"] == "Umum":
        json_data = common_table(dokumen, excel_path, hash_code)
    else:
        # json_data = uncommon_table(dokumen, excel_path, hash_code)
        json_data = ex_surat_penyerahan_barang(excel_path, hash_code)
    return json_data


# data transform for common table
def common_table(document, excel_path, hash_code):
    column_list = document["daftar_kolom"]
    # get list column list
    column_name_list = [column["nama_kolom"] for column in column_list]
    # get list idx column
    use_cols = [column["idx"] for column in column_list]
    # read excel
    df = pd.read_excel(
        excel_path, header=None, names=column_name_list, usecols=use_cols
    )
    # skiprows
    df[column_name_list[1]] = df[column_name_list[1]].fillna("")
    # skiprows = df[df[column_name_list[1]].str.contains("2")].index.tolist()[0] + 1
    skiprows = document["skiprows"]
    df = df.iloc[skiprows:]
    df = df.reset_index(drop=True)
    # replace "No" column value with generate number
    # df[column_name_list[0]] = range(1, len(df) + 1)

    # get list column with tipe_data int
    int_column_list = [
        column["nama_kolom"] for column in column_list if column["tipe_data"] == "int"
    ]
    # int_column_list = int_column_list[1:]
    # change NaN in int_column_list with 0 and convert to int
    for column_name in int_column_list:
        df[column_name] = df[column_name].fillna(0)
        convert_to_int(df, column_name)

    # transform to json
    json_filename = hash_code + ".json"
    json_path = os.path.join(JSON_DIR, json_filename)
    json_all_items = df_to_json(df, json_path)

    return json_all_items


# include check allowed_file and save image, return value : path image list
def files_handling(files):
    path_img_list = []
    for file in files:
        if allowed_file(file.filename):
            filename = secure_filename(file.filename)
            img_path = os.path.join(UPLOAD_IMAGE_PATH, filename)
            print(img_path)
            file.save(img_path)
            path_img_list.append(img_path)
    return path_img_list


# images stitching from list image
def images_stitching_1(img_path_list):
    img_list = []
    for path in img_path_list:
        img = cv2.imread(path)
        img_list.append(img)
        os.remove(path)

    # resize the width of the image to follow the first image
    for i in range(len(img_list) - 1):
        img_list[i + 1] = cv2.resize(
            img_list[i + 1], (img_list[0].shape[1], img_list[i + 1].shape[0])
        )

    # Create a blank image with sufficient size
    width = img_list[0].shape[1]
    height = 0
    for i in range(len(img_list)):
        height += img_list[i].shape[0]
    result = np.zeros((height, width, 3), dtype=np.uint8)

    # placement of images based on sequence
    start = 0
    for img in img_list:
        end = start + img.shape[0]
        result[start:end, :] = img
        start = end
    draw_line_result = draw_vertical_line(result)
    filename = "combined_image.jpg"
    file_path = os.path.join(UPLOAD_IMAGE_PATH, filename)

    # save image to file
    cv2.imwrite(file_path, draw_line_result)
    return file_path


def draw_vertical_line(image):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Mengaplikasikan Canny edge detection
    edges = cv2.Canny(gray, 50, 150)

    # Menggunakan Transformasi Hough untuk mendeteksi garis
    lines = cv2.HoughLinesP(
        edges, 1, np.pi / 180, threshold=100, minLineLength=100, maxLineGap=10
    )

    horizontal_lines = []
    vertical_lines = []
    for line in lines:
        x1, y1, x2, y2 = line[0]
        angle = np.arctan2(y2 - y1, x2 - x1) * 180 / np.pi

        if abs(angle) < 10:  # Menentukan ambang batas sudut garis horizontal
            horizontal_lines.append(line)
        elif abs(angle + 90) < 10:  # Menentukan ambang batas sudut garis vertikal
            vertical_lines.append(line)
        else:
            print(angle - 90)

    sorted_h_lines = sorted(
        horizontal_lines, key=lambda line: min(line[0][1], line[0][3]), reverse=True
    )
    ymax = sorted_h_lines[0][0][1]
    # Menampilkan garis vertikal
    for line in vertical_lines:
        x1, y1, x2, y2 = line[0]
        cv2.line(image, (x2, y2), (x1, ymax), (0, 0, 0), 1)
    return image


def images_stitching_2(img_path_list):
    image1 = cv2.imread(img_path_list[0])
    image2 = cv2.imread(img_path_list[1])
    # Konversi gambar ke skala abu-abu
    gray1 = cv2.cvtColor(image1, cv2.COLOR_BGR2GRAY)
    gray2 = cv2.cvtColor(image2, cv2.COLOR_BGR2GRAY)

    # Mendapatkan dimensi gambar pertama
    height1, width1 = gray1.shape[:2]

    # Hitung selisih tinggi antara gambar pertama dan gambar kedua
    height_diff = height1 - gray2.shape[0]

    # Tambahkan padding pada gambar kedua untuk menyamakan tinggi dengan gambar pertama
    padded_image2 = cv2.copyMakeBorder(
        image2, 0, height_diff, 0, 0, cv2.BORDER_CONSTANT, value=0
    )

    # Temukan pergeseran (shift) antara kedua gambar dengan template matching secara vertikal
    result = cv2.matchTemplate(gray1, gray2, cv2.TM_CCOEFF_NORMED)
    _, _, _, max_loc = cv2.minMaxLoc(result)
    y_shift = max_loc[1] - gray1.shape[0]

    # Gabungkan gambar pertama dengan gambar kedua yang telah di-shift secara vertikal
    stitched_image = np.concatenate((image1, padded_image2[y_shift:, :]), axis=0)

    # Gabungkan garis vertikal tabel dari gambar pertama dengan gambar kedua
    for y in range(height1):
        stitched_image[y, width1:] = np.where(
            stitched_image[y, width1:] == 255,
            image1[y, width1:],
            stitched_image[y, width1:],
        )

    height1, width1 = stitched_image.shape[:2]
    cropped_stitc = stitched_image[: height1 - height_diff, :]

    filename = "combined_image.jpg"
    file_path = os.path.join(UPLOAD_IMAGE_PATH, filename)

    # save image to file
    cv2.imwrite(file_path, cropped_stitc)
    return file_path


# transform dataframe to json
def df_to_json(df, json_path):
    json_str = df.to_json(orient="records")  # df to json string
    json_list = json.loads(json_str)  # str to json list
    with open(json_path, "w") as file:
        json.dump(json_list, file)
    return json_list


def convert_dict_to_int(id, data_dict):
    # get data by id, ambil nama dokumennya, abis itu dapetin data kolom dan tipe data nya
    pass


# convert each value in dataframe column to integer
def convert_to_int(df, column_name):
    print("Converting " + column_name + " to Int")
    for index, value in enumerate(df[column_name]):
        if type(df.at[index, column_name]) == str:
            if df.at[index, column_name].isnumeric() == False:
                df.at[index, column_name] = df.at[index, column_name].replace(".", "")
                df.at[index, column_name] = df.at[index, column_name].replace(",", "")
                if df.at[index, column_name].isnumeric() == False:
                    df.at[index, column_name] = 0
        df.at[index, column_name] = int(df.at[index, column_name])


# ga dipake
def ex_data_transform(dokument_type, excel_path, hash_code):
    json_data = []
    if dokument_type == "surat_penyerahan_barang":
        json_data = ex_surat_penyerahan_barang(
            excel_path=excel_path, hash_code=hash_code
        )
    elif dokument_type == "surat_penerimaan_materil":
        json_data = ex_surat_penerimaan_materil(
            excel_path=excel_path, hash_code=hash_code
        )
    elif dokument_type == "surat_kebutuhan_alat":
        json_data = ex_surat_kebutuhan_alat(excel_path=excel_path, hash_code=hash_code)
    return json_data


def uncommon_table(doc_type_data, excel_path, hash_code):
    # header
    column_list = doc_type_data["daftar_kolom"]
    column_name_list = [column["nama_kolom"] for column in column_list]
    # generate use cols: generate list 0..len(column name list)
    use_cols = list(range(len(column_name_list)))
    # read excel
    df = pd.read_excel(
        excel_path, header=None, names=column_name_list, usecols=use_cols
    )
    # skiprows
    nan_indices = df.index[df[column_name_list[1]].isna()].tolist()
    if nan_indices:
        print("adaan")
        df[column_name_list[1]] = df[column_name_list[1]].fillna("Null")
    skiprows = df[df[column_name_list[1]].str.contains("2")].index.tolist()[0] + 1
    if df.at[skiprows, column_name_list[1]] == "Null":
        print("masuk")
        skiprows += 1
    else:
        print("ga masuk")
    df = df.iloc[skiprows:]
    df = df.reset_index(drop=True)
    # get list column with tipe_data int and convert it (include change nan)
    int_column_list = [
        column["nama_kolom"] for column in column_list if column["tipe_data"] == "int"
    ]
    for column_name in int_column_list:
        df[column_name] = df[column_name].fillna(0)
        convert_to_int(df, column_name)
    # look for coordinates item separator
    item_separator_data = doc_type_data["pemisah_item"].copy()
    item_separator_data = add_data_type_column_to_dict(
        column_list, item_separator_data["kolom_patokan"], item_separator_data
    )
    # print(item_separator_data)
    list_idx_separator = keyword_handler(df, item_separator_data)
    # create new DataFrame for each item
    table_items = []
    start = list_idx_separator[0]
    for idx_item in range(1, len(list_idx_separator)):
        end = list_idx_separator[idx_item]
        item = df.iloc[start:end].reset_index(drop=True)
        table_items.append(item)
        start = end
    # Adds the last item
    last_item = df.iloc[start:, :]
    table_items.append(last_item)
    element_items = doc_type_data["elemen_dalam_satu_item"]
    print(element_items)

    json_all_items = []
    # for idx_item, item in enumerate(table_items):
    #     # divide item data into nama materill, seri, and kelengkapan
    #     # array_row_element = np.where(item["Nama Materil"].str.contains(":"))
    #     # looping untuk mendapatkan index elemen
    #     data_each_elements = []
    #     for element in element_items:
    #         #   dapetin index elemen dari suatu item
    #         #   dapetin tipe_data kolom  yang menjadi key column
    #         column_data = {}
    #         column_data["nama_kolom"] = element["key_column"]
    #         column_data["keyword"] = element["keyword"]
    #         column_data = add_data_type_column_to_dict(
    #             column_list, element["key_column"], column_data
    #         )
    #         idx_list = keyword_handler(item, column_data)
    #         # gimana jika dua elemen(bisa brurutan, bisa ada jeda) dari satu item, punya key kolom dan keyword yg sama
    #         # intinya keyword ny sama
    #         # gimana jika suatu item ngg punya elemen yg lengkap
    #         if idx_list:
    #             element_data = {}
    #             element_data["nama_elemen"] = element["nama_elemen"]
    #             element_data["tipe_data"] = element["tipe_data_elemen"]
    #             element_data["related_column"] = element["related_column"]
    #             element_data["indeks"] = idx_list[0]
    #             data_each_elements.append(element_data)

    #     # masukin "dataframe" ke data_each_elements
    #     # crop df item brdasarkan index dan related column
    #     if len(data_each_elements) == 1:
    #         if data_each_elements[0]["related_column"] == "Semua":
    #             element = item
    #         else:
    #             element = item.loc[
    #                 :, data_each_elements[0]["related_column"]
    #             ].reset_index(drop=True)

    #         data_each_elements[0]["dataframe"] = element
    #     elif len(data_each_elements) > 1:
    #         start = data_each_elements[0]["indeks"]
    #         for idx in range(1, len(data_each_elements)):
    #             end = data_each_elements[idx]["indeks"]
    #             rel_col = data_each_elements[idx - 1]["related_column"]
    #             if rel_col == "Semua":
    #                 element = item.loc[start:end].reset_index(drop=True)
    #             else:
    #                 element = item.loc[start:end, rel_col].reset_index(drop=True)
    #             data_each_elements[idx - 1]["dataframe"] = element
    #             start = end
    #         # Adds the last item
    #         rel_col = data_each_elements[-1]["related_column"]
    #         if rel_col == "Semua":
    #             element = item.loc[start:].reset_index(drop=True)
    #         else:
    #             element = item.loc[start:, rel_col].reset_index(drop=True)
    #         data_each_elements[-1]["dataframe"] = element

    #     # transform df berdasarkan tipe data elemen
    #     # pertama cek tipe data elemen, list atau json
    #     # klo json, bisa lngsung .to_json(orient="records")
    #     # abis itu ganti json string ke object
    #     # klo list, gitu lah
    #     json_item = [{}]
    #     for data in data_each_elements:
    #         if data["tipe_data"] == "json":
    #             json_string = data["dataframe"].to_json(orient="records")
    #             json_item[0][data["nama_elemen"]] = json.loads(json_string)
    #         elif data["tipe_data"] == "list":
    #             seri_str = ""
    #             for seri in data["dataframe"]:
    #                 seri_str += seri
    #             idx_seri = seri_str.find(":")
    #             seri_str = seri_str[idx_seri + 1 :]
    #             seri_list = seri_str.split(",")
    #             json_item[0][data["nama_elemen"]] = seri_list
    #         else:
    #             json_string = data["dataframe"].iloc[1:].to_json(orient="records")
    #             json_item[0][data["nama_elemen"]] = json.loads(json_string)
    #     json_all_items.append(json_item[0])
    # json_filename = hash_code + ".json"
    # json_path = os.path.join(JSON_DIR, json_filename)
    # with open(json_path, "w") as file:
    #     json.dump(json_all_items, file)
    return json_all_items


# tujuan : untuk menambahkan tipe data kolom ke suatu dict
def add_data_type_column_to_dict(column_list, searched_column_name, dict_needed_to_add):
    for column in column_list:
        if column["nama_kolom"] == searched_column_name:
            dict_needed_to_add["tipe_data"] = column["tipe_data"]
            break
    return dict_needed_to_add


# get index list that relate with keyword from dataframe
def keyword_handler(df, column_data):
    # index_list = []
    print(column_data)
    no_list = df[column_data["kolom_patokan"]].to_list()
    if column_data["tipe_data"] == "int":
        if column_data["keyword"] == "null":
            index_list = [i for i, value in enumerate(no_list) if value == 0]
            print("masuk null")
        # keyword == not null
        else:
            index_list = [i for i, value in enumerate(no_list) if value != 0]
            print("masuk !null")
    else:
        if column_data["keyword"] == "null":
            index_list = df[
                df[column_data["kolom_patokan"]].str.contains("Null")
            ].index.tolist()
        else:
            index_list = df[
                df[column_data["kolom_patokan"]].str.contains(column_data["keyword"])
            ].index.tolist()
    return index_list


def ex_surat_penyerahan_barang(excel_path, hash_code):
    header = [
        "No",
        "Nama Materil",
        "Jumlah (Angka)",
        "Jumlah (Huruf)",
        "Jumlah (Satuan)",
        "Keterangan",
    ]  # set header
    df = pd.read_excel(
        excel_path, header=None, skiprows=4, names=header, usecols=[0, 1, 2, 3, 4, 5]
    )  # read excel file
    df["No"] = df["No"].fillna(0).astype(int)  # change NaN to 0
    df["Jumlah (Angka)"] = df["Jumlah (Angka)"].fillna(0).astype(int)  # change NaN to 0
    list_row_item = np.where(
        df["Nama Materil"].isna()
    )  # look for coordinates that become boundaries between items

    # create new DataFrame for each item
    items_table = []
    start = 0
    for end in list_row_item[0]:
        item = df.iloc[start:end].reset_index(drop=True)
        items_table.append(item)
        start = end + 1

    # Adds the last item
    last_item = df.iloc[start:, :4]
    items_table.append(last_item)

    json_all_items = []

    # transform data for each item to json
    for i, item in enumerate(items_table):
        # divide item data into nama materill, seri, and kelengkapan
        array_row_element = np.where(item["Nama Materil"].str.contains(":"))

        # transform for data nama materill
        if len(array_row_element[0]) == 0:
            bound = len(item)
        else:
            bound = array_row_element[0][0]

        item_nama_materil = item.iloc[0:bound, :]
        json_materil = item_nama_materil.to_json(orient="records")
        json_item = json.loads(
            json_materil
        )  # change json_materil(string) to Python object (dict)

        # check if there are seri data
        if len(array_row_element[0]) >= 1:
            item_seri = item.iloc[
                array_row_element[0][0] : array_row_element[0][1], 1:2
            ]
            # transform for data seri
            seri_str = ""
            for data in item_seri["Nama Materil"]:
                seri_str += data
            index = seri_str.find(":")
            seri_str = seri_str[index + 1 :]
            seri_list = seri_str.split(",")
            json_item[0]["Seri"] = seri_list

        # check if there are kelengkapan data
        if len(array_row_element[0]) >= 2:
            item_kelengkapan = item.iloc[array_row_element[0][1] + 1 :, 1:]
            # transform for data kelengkapan
            json_kelengkapan = item_kelengkapan.to_json(orient="records")
            json_item[0]["Kelengkapan"] = json.loads(json_kelengkapan)

        # add json_item to json_all_item
        json_all_items.append(json_item[0])

    json_filename = hash_code + ".json"
    json_path = os.path.join(JSON_DIR, json_filename)
    with open(json_path, "w") as file:
        json.dump(json_all_items, file)
    return json_all_items


def ex_surat_penerimaan_materil(excel_path, hash_code):
    header = [
        "No Urut",
        "Nama dan Kode Materiil",
        "Satuan",
        "Banyaknya (Angka)",
        "Jumlah (Rp)",
        "Keterangan",
    ]
    df = pd.read_excel(
        excel_path, header=None, skiprows=3, names=header, usecols=[0, 1, 2, 3, 5, 6]
    )
    df["No Urut"] = range(0, len(df))

    # change NaN with 0
    df["Banyaknya (Angka)"] = df["Banyaknya (Angka)"].fillna(0)
    df["Jumlah (Rp)"] = df["Jumlah (Rp)"].fillna(0)
    # convert column data type to int
    column_to_convert = ["Banyaknya (Angka)", "Jumlah (Rp)"]
    for column_name in column_to_convert:
        convert_to_int(df, column_name)

    json_filename = hash_code + ".json"
    json_path = os.path.join(JSON_DIR, json_filename)
    json_all_items = df_to_json(df, json_path)

    return json_all_items


def ex_surat_kebutuhan_alat(excel_path, hash_code):
    header = [
        "No",
        "Jenis Materiil",
        "Satuan",
        "Indeks OPS",
        "Kebut OPS",
        "Nyata",
        "Terdukung",
        "Kurang",
        "Keterangan",
    ]
    df = pd.read_excel(
        excel_path,
        header=None,
        skiprows=2,
        names=header,
        usecols=[0, 1, 2, 3, 4, 5, 6, 7, 8],
    )
    df["No"] = range(1, len(df) + 1)

    # change NaN with 0
    df["Indeks OPS"] = df["Indeks OPS"].fillna(0)
    df["Kebut OPS"] = df["Kebut OPS"].fillna(0)
    df["Nyata"] = df["Nyata"].fillna(0)
    df["Terdukung"] = df["Terdukung"].fillna(0)
    df["Kurang"] = df["Kurang"].fillna(0)
    # convert column data type to int
    column_to_convert = ["Indeks OPS", "Kebut OPS", "Nyata", "Terdukung", "Kurang"]
    for column_name in column_to_convert:
        convert_to_int(df, column_name)

    json_filename = hash_code + ".json"
    json_path = os.path.join(JSON_DIR, json_filename)
    json_all_items = df_to_json(df, json_path)

    return json_all_items
